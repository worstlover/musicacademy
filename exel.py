import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def create_final_excel():
    """ساخت فایل اکسل نهایی - ۲ شیت ساده"""
    
    wb = openpyxl.Workbook()
    
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell_alignment = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    
    def write_headers(ws, headers, widths):
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
            cell.border = thin_border
        for i, width in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = width
    
    def write_row(ws, row_num, data):
        for col, value in enumerate(data, 1):
            cell = ws.cell(row=row_num, column=col, value=value)
            cell.alignment = cell_alignment
            cell.border = thin_border
    
    # ================= Sheet 1: هنرجویان =================
    ws1 = wb.active
    ws1.title = "هنرجویان"
    
    headers1 = [
        "ردیف",
        "نام",
        "نام خانوادگی",
        "کد ملی (اختیاری)",
        "شماره موبایل (اجباری)",
        "نام استاد (اجباری)",
        "نام کلاس/ساز (اختیاری - اگه خالی باشد از تخصص استاد می‌خوانیم)",
        "نام والدین (اختیاری)",
        "شماره والدین (اختیاری)",
        "تاریخ تولد (اختیاری)",
        "آدرس (اختیاری)",
        "مانده حساب اولیه (تومان)"
    ]
    
    widths1 = [6, 12, 12, 15, 18, 15, 20, 15, 15, 15, 25, 18]
    write_headers(ws1, headers1, widths1)
    
    # نمونه‌ها
    samples1 = [
        [1, "علی", "محمدی", "0012345678", "09123456789", "مریم حسینی", "پیانو", "", "", "", "", 500000],
        [2, "زهرا", "رضایی", "0023456789", "09351234567", "مریم حسینی", "پیانو", "", "", "1385/03/15", "", 300000],
        [3, "حسین", "کریمی", "", "09124567890", "رضا کریمی", "", "", "", "", "", 0],
    ]
    
    for i, sample in enumerate(samples1, 2):
        write_row(ws1, i, sample)
    
    # راهنما
    ws1.cell(row=6, column=1, value="راهنما:").font = Font(bold=True, size=13)
    ws1.cell(row=7, column=1, value="• شماره موبایل: ۱۱ رقم، با ۰۹ (مثال: 09123456789)").font = Font(italic=True)
    ws1.cell(row=8, column=1, value="• نام استاد: باید دقیقاً با شیت «اساتید» یکی باشد").font = Font(italic=True)
    ws1.cell(row=9, column=1, value="• نام کلاس/ساز: اگه خالی باشد، از تخصص استاد استفاده می‌شود").font = Font(italic=True)
    ws1.cell(row=10, column=1, value="• مانده حساب: پولی که هنرجو الان دارد (اگه بدهکار است عدد منفی)").font = Font(italic=True)
    ws1.cell(row=11, column=1, value="• یک هنرجو می‌تواند چند ردیف داشته باشد (برای چند استاد)").font = Font(bold=True, color="FF0000")
    ws1.cell(row=12, column=1, value="  مثال: علی محمدی ۲ ردیف دارد - یکی با مریم حسینی، یکی با رضا کریمی").font = Font(italic=True)
    
    # ================= Sheet 2: اساتید =================
    ws2 = wb.create_sheet("اساتید")
    
    headers2 = [
        "ردیف",
        "نام",
        "نام خانوادگی",
        "کد ملی (اختیاری)",
        "شماره موبایل (اجباری)",
        "تخصص/ساز",
        "تعرفه ساعتی انفرادی (تومان)",
        "تعرفه ساعتی گروهی (تومان)",
        "مانده حساب استاد (تومان)"
    ]
    
    widths2 = [6, 12, 12, 15, 18, 20, 22, 22, 22]
    write_headers(ws2, headers2, widths2)
    
    samples2 = [
        [1, "مریم", "حسینی", "0098765432", "09351234567", "پیانو", 150000, 200000, 0],
        [2, "رضا", "کریمی", "0087654321", "09123456789", "گیتار", 120000, 180000, 500000],
        [3, "سارا", "احمدی", "", "09198765432", "ویلن", 180000, 220000, 0],
    ]
    
    for i, sample in enumerate(samples2, 2):
        write_row(ws2, i, sample)
    
    ws2.cell(row=6, column=1, value="راهنما:").font = Font(bold=True, size=13)
    ws2.cell(row=7, column=1, value="• تعرفه انفرادی: هزینه ۱ ساعت کلاس خصوصی").font = Font(italic=True)
    ws2.cell(row=8, column=1, value="• تعرفه گروهی: هزینه ۱ ساعت کلاس گروهی").font = Font(italic=True)
    ws2.cell(row=9, column=1, value="• مانده حساب: بدهی آموزشگاه به استاد (اگه استاد بدهکاره عدد منفی)").font = Font(italic=True)
    ws2.cell(row=10, column=1, value="• تخصص/ساز: مثلاً پیانو، گیتار، ویلن، سنتور، آواز...").font = Font(italic=True)
    
    wb.save("music_academy_data.xlsx")
    print("✅ فایل ساخته شد!")
    print("\n📋 شیت‌ها:")
    print("   1. هنرجویان (با نام استاد)")
    print("   2. اساتید")
    print("\n📌 فقط همین ۲ شیت رو پر کنید.")
    print("📌 اگه هنرجو ۲ استاد داره، ۲ ردیف برایش بسازید.")

create_final_excel()