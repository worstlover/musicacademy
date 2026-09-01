import openpyxl
from core.models import Student, Teacher, Course, StudentCourse, RateTemplate, TeacherRate, WalletTransaction
import jdatetime
from datetime import datetime

def import_from_excel(file_path):
    """
    import داده‌ها از فایل اکسل
    """
    wb = openpyxl.load_workbook(file_path)
    results = {
        'rate_templates': 0,
        'teachers': 0,
        'teacher_rates': 0,
        'students': 0,
        'courses': 0,
        'enrollments': 0,
        'errors': [],
    }
    
    # ================= Sheet 1: قالب‌های تعرفه =================
    if 'قالب‌های تعرفه' in wb.sheetnames:
        ws = wb['قالب‌های تعرفه']
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row[0]:  # اگه ردیف خالی بود
                continue
            try:
                rate, created = RateTemplate.objects.get_or_create(
                    name=row[1].strip(),
                    defaults={'description': row[2] if len(row) > 2 else None}
                )
                if created:
                    results['rate_templates'] += 1
            except Exception as e:
                results['errors'].append(f"RateTemplate error: {e}")
    
    # ================= Sheet 2: اساتید =================
    if 'اساتید' in wb.sheetnames:
        ws = wb['اساتید']
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row[0]:
                continue
            try:
                teacher, created = Teacher.objects.get_or_create(
                    phone_number=str(row[3]).strip(),
                    defaults={
                        'first_name': row[1],
                        'last_name': row[2],
                        'specialization': row[4] if len(row) > 4 else '',
                        'commission_percent': int(row[5]) if len(row) > 5 and row[5] else 30,
                        'initial_balance': int(row[6]) if len(row) > 6 and row[6] else 0,
                    }
                )
                if created:
                    results['teachers'] += 1
            except Exception as e:
                results['errors'].append(f"Teacher error: {e}")
    
    # ================= Sheet 3: تعرفه اساتید =================
    if 'تعرفه اساتید' in wb.sheetnames:
        ws = wb['تعرفه اساتید']
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row[0]:
                continue
            try:
                teacher = Teacher.objects.filter(
                    first_name=row[1], last_name=row[2]
                ).first()
                rate_template = RateTemplate.objects.filter(name=row[3]).first()
                
                if teacher and rate_template:
                    tr, created = TeacherRate.objects.get_or_create(
                        teacher=teacher,
                        rate_template=rate_template,
                        defaults={'hourly_rate': int(row[4])}
                    )
                    if created:
                        results['teacher_rates'] += 1
            except Exception as e:
                results['errors'].append(f"TeacherRate error: {e}")
    
    # ================= Sheet 4: هنرجویان =================
    if 'هنرجویان' in wb.sheetnames:
        ws = wb['هنرجویان']
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row[0]:
                continue
            try:
                phone = str(row[3]).strip()
                student, created = Student.objects.get_or_create(
                    phone_number=phone,
                    defaults={
                        'first_name': row[1],
                        'last_name': row[2],
                        'national_code': row[4] if len(row) > 4 else None,
                    }
                )
                
                if created:
                    results['students'] += 1
                    
                    # ثبت مانده حساب اولیه
                    initial_balance = int(row[7]) if len(row) > 7 and row[7] else 0
                    if initial_balance > 0:
                        WalletTransaction.objects.create(
                            student=student,
                            transaction_type='credit',
                            amount=initial_balance,
                            description="مانده حساب اولیه",
                            status='approved',
                            payment_method='manual'
                        )
                    
                    # اتصال به استاد
                    if len(row) > 5 and row[5]:
                        teacher = Teacher.objects.filter(
                            first_name=row[5].split()[0],
                            last_name=row[5].split()[-1] if len(row[5].split()) > 1 else ''
                        ).first()
                        
                        if teacher:
                            # انتخاب تعرفه
                            rate_name = row[6] if len(row) > 6 and row[6] else teacher.specialization
                            rate_template = RateTemplate.objects.filter(name=rate_name).first()
                            
                            course, _ = Course.objects.get_or_create(
                                name=f"{teacher.specialization} - {rate_template.name if rate_template else 'پایه'}",
                                teacher=teacher,
                                rate_template=rate_template,
                                defaults={
                                    'duration_minutes': 60,
                                    'base_fee': 100000,
                                }
                            )
                            
                            StudentCourse.objects.create(student=student, course=course)
                            results['enrollments'] += 1
                            
            except Exception as e:
                results['errors'].append(f"Student error row {row[0]}: {e}")
    
    return results